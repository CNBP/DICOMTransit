import sys
import os


from pathlib import Path

path_file = os.path.dirname(os.path.realpath(__file__))
path_module = Path(path_file).parents[
    1
]  # 0 is /DICOMTransit/DICOMTransit, 1 is /DICOMTransit

sys.path.append(f"{path_module}/Python/")

# print(sys.path)
# This task is trigger via a CRON job.

from DICOMTransit.LocalDB.API import get_list_MRN, get_list_CNBPID
from openpyxl import Workbook
from PythonUtils.PUFile import unique_name


# This scheduler script runs a daily export of the the list CNBPID and MRN correspondence
print("Aggregating information from DICOMTransit database")
list_MRN = get_list_MRN()
list_CNBPID = get_list_CNBPID()
list_correspondence = zip(list_MRN, list_CNBPID)
print(f"Total entry of {len(list_MRN)} subjects found from the DICOMTransit database")


wb = Workbook()

ws = wb.active
for row_index, (key, value) in enumerate(list_correspondence):
    ws.cell(column=1, row=row_index + 1, value=key)
    ws.cell(column=2, row=row_index + 1, value=value)

# Set basic file modifying password to prevent accidental editing of the correspondence
wb.security.workbookPassword = "readonly"
wb.security.lockStructure = True

# save files.
# fixme: this should be a user configurable variable.
detination_filename = f"/toshiba2/PRO_001_CNBP_LORIS_Archive/CNBPID_correspondence/CNBPID_list_{unique_name()}.xlsx"
print(f"Saveing XLSX to {detination_filename}")
wb.save(detination_filename)
